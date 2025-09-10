# -*- coding: utf-8 -*-
# avstamning_master.py – K1–K6 + K4N, GroupKey-stämpling vid match, robust Excel & subset-sum

import re, math, itertools, warnings, zipfile
from itertools import count
from pathlib import Path
from typing import Optional, List, Set

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore", category=UserWarning, module=r"openpyxl\.styles\.stylesheet")

# ===================== KONSTANTER =====================
BANK_COLS = [
    "Bokföringsdatum","Valutadatum","Referens","Text","Motkonto","Belopp",
    "Medgivandereferens","Betalningsmottagarens identitet","Transaktionskod"
]
BOKF_COLS = [
    "Gruppering: (KTO-ANS-SPE)","FTG","KTO","SPE","ANS","OBJ","MOT",
    "PRD","MAR","RGR","Datum","IB Året SEK","Ing. ack. belopp 07-2025 SEK",
    "Period SEK","Utg. ack. belopp 07-2025 SEK","Val","Utländskt valutabelopap",
    "Text1","Postning -Dokumentsekvensnummer","Verifikationsnummer","Källa","Kategori"
]
KOMB_COLS = [
    "Gruppering: (KTO-ANS-SPE)","FTG","KTO","SPE","ANS","OBJ","MOT","PRD","MAR","RGR",
    "Datum","IB Året SEK","Ing. ack. Belopp","Period SEK","Utg. ack. Belopp","Val",
    "Utländskt valutabelopp","Text","Postning -Dokumentsekvensnummer","Verifikationsnummer",
    "Källa","Kategori","System","Ny källa","MatchKategori","MatchGruppID",
]
BANK_HEADER_ROW = 4
BOKF_HEADER_ROW = 17
AMOUNT_ALIASES = ["Period SEK", "Period\u00A0SEK", "Period  SEK", "PeriodSEK"]

# Global sekvens för GroupKey
_GROUP_SEQ = count(1)

# ===================== DIALOGER =====================
def ask_file_dialog(title="Välj fil") -> Optional[str]:
    try:
        import tkinter as tk
        from tkinter import filedialog, messagebox
        root = tk.Tk(); root.withdraw()
        messagebox.showinfo("Välj fil", title)
        path = filedialog.askopenfilename(title=title, filetypes=[("Excel/CSV","*.xlsx *.xls *.csv")])
        root.destroy()
        return path
    except Exception:
        return None

def pick_file_with_validation(kind: str) -> str:
    while True:
        title = "Välj kontoutdraget" if kind == "Bank" else "Välj bokföringslistan"
        path = ask_file_dialog(title)
        if not path:
            path = input(f"Sökväg till {kind}-fil: ").strip()
        try:
            _ = load_bank(path) if kind == "Bank" else load_bokf(path)
            return path
        except Exception as e:
            print(f"\n❗ Fel fil för {kind}: {e}\nFörsök igen.\n")

def ask_save_as_dialog(default_name="output_avstamning.xlsx", initialdir=None) -> Optional[str]:
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk(); root.withdraw()
        path = filedialog.asksaveasfilename(
            title="Välj var resultatfilen ska sparas",
            defaultextension=".xlsx", initialfile=default_name, initialdir=initialdir,
            filetypes=[("Excel-fil","*.xlsx")]
        )
        root.destroy()
        return path
    except Exception:
        return None

# ===================== STÄDNING & KOLUMN-NORMALISERING =====================
_XML_BAD = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")

def sanitize_xml(v):
    if v is None or (isinstance(v, float) and pd.isna(v)): return ""
    s = str(v)
    s = _XML_BAD.sub("", s)
    return s

def sanitize_all_text_cols(df: pd.DataFrame) -> pd.DataFrame:
    str_cols = df.select_dtypes(include=["object"]).columns
    for c in str_cols:
        df[c] = df[c].apply(sanitize_xml)
    return df

def clean_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    df = sanitize_all_text_cols(df.copy())
    for c in df.columns:
        if pd.api.types.is_object_dtype(df[c]):
            df[c] = df[c].fillna("")
    return df

def normalize_column_names(df: pd.DataFrame) -> pd.DataFrame:
    new_cols = []
    for c in df.columns:
        c2 = str(c).replace("\u00A0", " ")
        c2 = re.sub(r"\s+", " ", c2).strip()
        new_cols.append(c2)
    df.columns = new_cols
    alias_map = {"Period  SEK": "Period SEK", "PeriodSEK": "Period SEK"}
    df.rename(columns=alias_map, inplace=True)
    return df

def get_amount_series(df: pd.DataFrame) -> pd.Series:
    df = normalize_column_names(df)
    for name in AMOUNT_ALIASES:
        if name in df.columns:
            return pd.to_numeric(df[name], errors="coerce")
    raise KeyError("Kunde inte hitta beloppskolumn (t.ex. 'Period SEK') i aktuellt urval.")

# ===================== TYPER & SUMMERING =====================
def _to_float(series: pd.Series) -> pd.Series:
    s = (series.astype(str)
         .str.replace(" ", "", regex=False)
         .str.replace("\u00a0", "", regex=False)
         .str.replace(",", ".", regex=False))
    return pd.to_numeric(s, errors="coerce")

def _strip_df(df: pd.DataFrame) -> pd.DataFrame:
    for c in df.columns:
        if pd.api.types.is_string_dtype(df[c]):
            df[c] = df[c].astype(str).str.strip()
    return df

def sek_round(x): return round(float(x), 2) if pd.notna(x) else x
def sum_sek(s): return sek_round(s.fillna(0).sum())
def startswith_seb(v): return isinstance(v,str) and v.upper().startswith("SEB")
def extract_yymmdd(dt):
    if pd.isna(dt): return None
    return pd.to_datetime(dt).strftime("%y%m%d")
def has_yymmdd_in_text1(t, y): return isinstance(t,str) and bool(re.search(rf"(?<!\d){re.escape(y)}(?!\d)", t))
def has_yymmdd_in_vnr(v, y):   return isinstance(v,str) and bool(re.search(rf"(?<!\d){re.escape(y)}(?!\d)", v))
def is_6digit_vnr(v):          return isinstance(v,str) and len(re.findall(r"(?<!\d)\d{6}(?!\d)", v or ""))>0

def col_apply(df: pd.DataFrame, col: str, func) -> pd.Series:
    if col in df.columns:
        return df[col].apply(func)
    return pd.Series([False]*len(df), index=df.index)

# ===================== INLÄSNING =====================
def load_bank(path: str) -> pd.DataFrame:
    p = Path(path)
    if p.suffix.lower() in [".xlsx",".xls"]:
        df = pd.read_excel(p, header=BANK_HEADER_ROW, dtype=str)
    else:
        df = pd.read_csv(p, skiprows=BANK_HEADER_ROW, dtype=str, sep=None, engine="python")
    for col in ["Bokföringsdatum","Text","Belopp"]:
        if col not in df.columns:
            raise ValueError(f"Bankfilen saknar kolumnen: '{col}'")
    df = normalize_column_names(_strip_df(df))
    df["Bokföringsdatum"] = pd.to_datetime(df["Bokföringsdatum"], errors="coerce")
    df["Belopp"] = _to_float(df["Belopp"])
    df = sanitize_all_text_cols(df)
    df = df.reset_index(drop=False).rename(columns={"index":"BankRowID"})
    return df

def load_bokf(path: str) -> pd.DataFrame:
    p = Path(path)
    if p.suffix.lower() in [".xlsx",".xls"]:
        df = pd.read_excel(p, header=BOKF_HEADER_ROW, dtype=str)
    else:
        df = pd.read_csv(p, skiprows=BOKF_HEADER_ROW, dtype=str, sep=None, engine="python")
    df = normalize_column_names(df)
    for col in ["Datum","IB Året SEK","Text1","Verifikationsnummer","Kategori"]:
        if col not in df.columns:
            raise ValueError(f"Bokföringsfilen saknar kolumnen: '{col}'")
    df = _strip_df(df)
    df = df[df["IB Året SEK"].isna() | (df["IB Året SEK"] == "")].copy()
    df["Datum"] = pd.to_datetime(df["Datum"], errors="coerce")
    if "Period SEK" in df.columns:
        df["Period SEK"] = _to_float(df["Period SEK"])
    else:
        found = next((nm for nm in AMOUNT_ALIASES if nm in df.columns), None)
        if found:
            df["Period SEK"] = _to_float(df[found])
        else:
            raise ValueError("Ingen beloppskolumn (Period SEK) hittad i bokföringsfilen.")
    df = sanitize_all_text_cols(df)
    df = df.reset_index(drop=False).rename(columns={"index":"BokfRowID"})
    return df

# ===================== SUBSET-SUM (meet-in-the-middle) =====================
def subset_sum_meet_in_middle(indices: List[int], amounts: List[float], target: float, tol: float=0.005) -> Optional[Set[int]]:
    if not indices: return None
    cents = [int(round((a if pd.notna(a) else 0.0) * 100)) for a in amounts]
    tgt = int(round(target * 100))
    n = len(indices); mid = n // 2
    left = list(zip(indices[:mid], cents[:mid])); right = list(zip(indices[mid:], cents[mid:]))

    def all_subset_sums(items):
        res = []; m = len(items)
        for mask in range(1<<m):
            s = 0; pick = set()
            for i in range(m):
                if mask & (1<<i):
                    s += items[i][1]; pick.add(items[i][0])
            res.append((s, pick))
        return res

    L = all_subset_sums(left); R = all_subset_sums(right)
    R.sort(key=lambda x: x[0]); import bisect
    r_sums = [x[0] for x in R]
    band = int(round(tol*100))
    for ls, lpick in L:
        need = tgt - ls
        i = bisect.bisect_left(r_sums, need - band)
        while i < len(R) and R[i][0] <= need + band:
            rs, rpick = R[i]
            if abs((ls+rs) - tgt) <= band:
                return set(lpick | rpick)
            i += 1
    return None

def drop_subset_if_sum_matches(df: pd.DataFrame, target_sum: float) -> Optional[pd.DataFrame]:
    if df.empty: return None
    amt = get_amount_series(df)
    base_sum = sum_sek(amt)
    remove_needed = sek_round(base_sum - target_sum)
    if math.isclose(remove_needed, 0.0, abs_tol=0.005):
        return df.copy()
    idx = df.index.tolist()
    amounts = amt.astype(float).tolist()
    hit = subset_sum_meet_in_middle(idx, amounts, remove_needed, tol=0.005)
    if hit is None: return None
    return df.drop(index=list(hit))

# ===================== GROUPKEY-HJÄLPARE =====================
def make_group_key(cat: str, bank_rows: pd.DataFrame) -> str:
    """Format: <KAT>-B<minBankRowID>-<löpnummer> (datumoberoende och unikt per körning)"""
    bid = 0 if bank_rows is None or bank_rows.empty else int(bank_rows["BankRowID"].min())
    n = next(_GROUP_SEQ)
    return f"{cat}-B{bid}-{n:06d}"

# ===================== K1 =====================
def run_category1_BG04803458(bank_df, bokf_df):
    bank_k1 = bank_df[
        bank_df["Text"].astype(str).str.contains(r"BG04803458", case=False, na=False)
        & (bank_df["Belopp"] > 0)
    ].copy()

    matched_bank_all, matched_bokf_all, used_bokf = [], [], set()

    def stamp(b_rows: pd.DataFrame, f_rows: pd.DataFrame):
        gk = make_group_key("K1", b_rows)
        b = b_rows.copy(); f = f_rows.copy()
        b["__MatchKategori__"]="K1"; b["__GroupKey__"]=gk
        f["__MatchKategori__"]="K1"; f["__GroupKey__"]=gk
        matched_bank_all.append(b); matched_bokf_all.append(f); return set(f["BokfRowID"])

    for _, bank_day_rows in bank_k1.groupby(bank_k1["Bokföringsdatum"].dt.date):
        bank_day_rows = bank_day_rows.sort_values("BankRowID")
        bank_sum = sum_sek(bank_day_rows["Belopp"])
        yymmdd = extract_yymmdd(bank_day_rows["Bokföringsdatum"].iloc[0])

        bokf_day = bokf_df[
            (bokf_df["Datum"].dt.date.isin(bank_day_rows["Bokföringsdatum"].dt.date.unique())) &
            (bokf_df["Kategori"].astype(str).str.strip().str.lower() == "inbetalningar") &
            (bokf_df["Period SEK"] > 0) & (~bokf_df["BokfRowID"].isin(used_bokf))
        ].copy()
        if bokf_day.empty: continue
        try_match = lambda df_now: math.isclose(sum_sek(get_amount_series(df_now)), bank_sum, abs_tol=0.005)

        # 1 total
        cur = bokf_day.copy()
        if try_match(cur): used_bokf |= stamp(bank_day_rows, cur); continue
        # 2 drop EN
        cur = bokf_day.copy()
        diff = sek_round(sum_sek(get_amount_series(cur)) - bank_sum)
        if diff != 0:
            amt = get_amount_series(cur)
            cand = cur[amt.round(2)==diff]
            if not cand.empty:
                cur2 = cur[cur["BokfRowID"]!=cand.iloc[0]["BokfRowID"]]
                if try_match(cur2): used_bokf |= stamp(bank_day_rows, cur2); continue
        # 3 SEB only
        cur = bokf_day[col_apply(bokf_day, "Verifikationsnummer", startswith_seb)].copy()
        if not cur.empty and try_match(cur): used_bokf |= stamp(bank_day_rows, cur); continue
        # 4 SEB + drop EN
        cur = bokf_day[col_apply(bokf_day, "Verifikationsnummer", startswith_seb)].copy()
        if not cur.empty:
            diff = sek_round(sum_sek(get_amount_series(cur)) - bank_sum)
            if diff != 0:
                amt = get_amount_series(cur)
                cand = cur[amt.round(2)==diff]
                if not cand.empty:
                    cur2 = cur[cur["BokfRowID"]!=cand.iloc[0]["BokfRowID"]]
                    if try_match(cur2): used_bokf |= stamp(bank_day_rows, cur2); continue
        # 5 icke-SEB subset-sum
        cur = bokf_day.copy()
        cur2 = drop_subset_if_sum_matches(cur, bank_sum)
        if cur2 is not None and try_match(cur2): used_bokf |= stamp(bank_day_rows, cur2); continue
        # 6 SEB + icke-SEB med rätt YYMMDD
        cur_all = bokf_day.copy()
        nonseb = ~col_apply(cur_all, "Verifikationsnummer", startswith_seb)
        right = col_apply(cur_all, "Verifikationsnummer", lambda v: has_yymmdd_in_vnr(v, yymmdd))
        non_seb_right = cur_all[nonseb & right]
        cur = pd.concat([cur_all[col_apply(cur_all,"Verifikationsnummer",startswith_seb)], non_seb_right])
        if not cur.empty and try_match(cur): used_bokf |= stamp(bank_day_rows, cur); continue
        # 7 rätt YYMMDD + drop EN
        if not cur.empty:
            diff = sek_round(sum_sek(get_amount_series(cur)) - bank_sum)
            if diff != 0:
                amt = get_amount_series(cur)
                cand = cur[amt.round(2)==diff]
                if not cand.empty:
                    cur2 = cur[cur["BokfRowID"]!=cand.iloc[0]["BokfRowID"]]
                    if try_match(cur2): used_bokf |= stamp(bank_day_rows, cur2); continue
        # 8 rätt YYMMDD + subset-sum
        if not cur.empty:
            cur2 = drop_subset_if_sum_matches(cur, bank_sum)
            if cur2 is not None and try_match(cur2): used_bokf |= stamp(bank_day_rows, cur2); continue

    matched_bank = pd.concat(matched_bank_all, ignore_index=True) if matched_bank_all else bank_k1.iloc[0:0].copy()
    matched_bokf = pd.concat(matched_bokf_all, ignore_index=True) if matched_bokf_all else bokf_df.iloc[0:0].copy()
    return matched_bank, matched_bokf

# ===================== K2 (BG04868550; YYMMDD i Text1/Vnr) =====================
def run_category2_BG04868550(bank_df, bokf_df):
    bank_k2 = bank_df[
        bank_df["Text"].astype(str).str.match(r"^\s*BG04868550", case=False, na=False)
        & (bank_df["Belopp"] > 0)
    ].copy()
    matched_bank_all, matched_bokf_all, used_bokf = [], [], set()

    def stamp(b_rows: pd.DataFrame, f_rows: pd.DataFrame):
        gk = make_group_key("K2", b_rows)
        b = b_rows.copy(); f = f_rows.copy()
        b["__MatchKategori__"]="K2"; b["__GroupKey__"]=gk
        f["__MatchKategori__"]="K2"; f["__GroupKey__"]=gk
        matched_bank_all.append(b); matched_bokf_all.append(f); return set(f["BokfRowID"])

    for bank_date, bank_day_rows in bank_k2.groupby(bank_k2["Bokföringsdatum"].dt.date):
        bank_day_rows = bank_day_rows.sort_values("BankRowID")
        bank_sum = sum_sek(bank_day_rows["Belopp"])
        yymmdd = extract_yymmdd(pd.to_datetime(bank_date))

        def base_A():
            base = bokf_df[
                (~bokf_df["BokfRowID"].isin(used_bokf)) &
                (bokf_df["Kategori"].astype(str).str.strip().isin(["SkA Bank","Inköpsfakturor"])) &
                (bokf_df["Period SEK"] > 0)
            ].copy()
            mask_text1 = col_apply(base, "Text1", lambda t: has_yymmdd_in_text1(t, yymmdd))
            return base[mask_text1].copy()

        def add_INB(df):
            extra = bokf_df[
                (~bokf_df["BokfRowID"].isin(used_bokf)) &
                (bokf_df["Kategori"].astype(str).str.strip()=="Inbetalningar") &
                (bokf_df["Period SEK"] > 0) &
                (col_apply(bokf_df, "Verifikationsnummer", lambda v: has_yymmdd_in_vnr(v, yymmdd)))
            ].copy()
            return pd.concat([df, extra], ignore_index=False)

        def add_BET(df):
            extra = bokf_df[
                (~bokf_df["BokfRowID"].isin(used_bokf)) &
                (bokf_df["Kategori"].astype(str).str.strip()=="Betalningar") &
                (bokf_df["Period SEK"] > 0) &
                (col_apply(bokf_df, "Verifikationsnummer", lambda v: has_yymmdd_in_vnr(v, yymmdd)))
            ].copy()
            return pd.concat([df, extra], ignore_index=False)

        try_match = lambda df_now: math.isclose(sum_sek(get_amount_series(df_now)), bank_sum, abs_tol=0.005)

        # A
        cur = base_A()
        if not cur.empty and try_match(cur): used_bokf |= stamp(bank_day_rows, cur); continue
        cur = base_A()
        if not cur.empty:
            amt = get_amount_series(cur); cand = cur[amt.round(2)==bank_sum]
            if not cand.empty:
                chosen = cand.iloc[[0]]; used_bokf |= stamp(bank_day_rows, chosen); continue
        cur = base_A()
        if not cur.empty:
            diff = sek_round(sum_sek(get_amount_series(cur)) - bank_sum)
            if diff != 0:
                amt=get_amount_series(cur); drop = cur[amt.round(2)==diff]
                if not drop.empty:
                    cur2 = cur[cur["BokfRowID"]!=drop.iloc[0]["BokfRowID"]]
                    if try_match(cur2): used_bokf |= stamp(bank_day_rows, cur2); continue
        cur = base_A()
        if not cur.empty:
            cur2 = drop_subset_if_sum_matches(cur, bank_sum)
            if cur2 is not None and try_match(cur2): used_bokf |= stamp(bank_day_rows, cur2); continue

        # A + INB
        cur = add_INB(base_A())
        if not cur.empty and try_match(cur): used_bokf |= stamp(bank_day_rows, cur); continue
        if not cur.empty:
            amt=get_amount_series(cur); cand = cur[amt.round(2)==bank_sum]
            if not cand.empty:
                chosen = cand.iloc[[0]]; used_bokf |= stamp(bank_day_rows, chosen); continue
        if not cur.empty:
            diff = sek_round(sum_sek(get_amount_series(cur)) - bank_sum)
            if diff != 0:
                amt=get_amount_series(cur); drop = cur[amt.round(2)==diff]
                if not drop.empty:
                    cur2 = cur[cur["BokfRowID"]!=drop.iloc[0]["BokfRowID"]]
                    if try_match(cur2): used_bokf |= stamp(bank_day_rows, cur2); continue
        if not cur.empty:
            cur2 = drop_subset_if_sum_matches(cur, bank_sum)
            if cur2 is not None and try_match(cur2): used_bokf |= stamp(bank_day_rows, cur2); continue

        # A + INB + BET
        cur = add_BET(add_INB(base_A()))
        if not cur.empty and try_match(cur): used_bokf |= stamp(bank_day_rows, cur); continue
        if not cur.empty:
            amt=get_amount_series(cur); cand = cur[amt.round(2)==bank_sum]
            if not cand.empty:
                chosen=cand.iloc[[0]]; used_bokf |= stamp(bank_day_rows, chosen); continue
        if not cur.empty:
            diff = sek_round(sum_sek(get_amount_series(cur)) - bank_sum)
            if diff != 0:
                amt=get_amount_series(cur); drop = cur[amt.round(2)==diff]
                if not drop.empty:
                    cur2 = cur[cur["BokfRowID"]!=drop.iloc[0]["BokfRowID"]]
                    if try_match(cur2): used_bokf |= stamp(bank_day_rows, cur2); continue
        if not cur.empty:
            cur2 = drop_subset_if_sum_matches(cur, bank_sum)
            if cur2 is not None and try_match(cur2): used_bokf |= stamp(bank_day_rows, cur2); continue

    matched_bank = pd.concat(matched_bank_all, ignore_index=True) if matched_bank_all else bank_k2.iloc[0:0].copy()
    matched_bokf = pd.concat(matched_bokf_all, ignore_index=True) if matched_bokf_all else bokf_df.iloc[0:0].copy()
    return matched_bank, matched_bokf

# ===================== K3 (35-referens 1:1) =====================
def run_category3_35ref(bank_df, bokf_df):
    has_35ref = bank_df["Text"].astype(str).str.contains(r"35\d{10}", regex=True, na=False)
    bank_k3 = bank_df[has_35ref].copy().sort_values(["Bokföringsdatum","BankRowID"])
    bokf_pay = bokf_df[(bokf_df["Kategori"].astype(str).str.strip() == "Betalningar")].copy()

    matched_bank_all, matched_bokf_all, used = [], [], set()

    def stamp(b_row: pd.DataFrame, f_row: pd.DataFrame):
        gk = make_group_key("K3", b_row)
        b = b_row.copy(); f = f_row.copy()
        b["__MatchKategori__"]="K3"; b["__GroupKey__"]=gk
        f["__MatchKategori__"]="K3"; f["__GroupKey__"]=gk
        matched_bank_all.append(b); matched_bokf_all.append(f)

    for _, b in bank_k3.iterrows():
        b_date = pd.to_datetime(b["Bokföringsdatum"]).date() if pd.notna(b["Bokföringsdatum"]) else None
        amount = sek_round(b["Belopp"])
        if b_date is None or pd.isna(amount): continue
        amt = get_amount_series(bokf_pay)
        cand = bokf_pay[(bokf_pay["Datum"].dt.date == b_date) & (~bokf_pay["BokfRowID"].isin(used)) & (amt.round(2) == amount)].copy()
        if len(cand) >= 1:
            chosen = cand.sort_values("BokfRowID").iloc[[0]]
            used |= set(chosen["BokfRowID"])
            stamp(b.to_frame().T, chosen)

    matched_bank = pd.concat(matched_bank_all, ignore_index=True) if matched_bank_all else bank_k3.iloc[0:0].copy()
    matched_bokf = pd.concat(matched_bokf_all, ignore_index=True) if matched_bokf_all else bokf_df.iloc[0:0].copy()
    return matched_bank, matched_bokf

# ===================== K4 (Övrigt 1:1) =====================
def run_category4_ovrigt(bank_df, bokf_df):
    mask_k1 = bank_df["Text"].astype(str).str.contains(r"BG04803458", case=False, na=False)
    mask_k2 = bank_df["Text"].astype(str).str.match(r"^\s*BG04868550", case=False, na=False)
    mask_k3 = bank_df["Text"].astype(str).str.contains(r"35\d{10}", regex=True, na=False)
    bank_k4 = bank_df[~(mask_k1 | mask_k2 | mask_k3)].copy().sort_values(["Bokföringsdatum","BankRowID"])

    matched_bank_all, matched_bokf_all, used = [], [], set()

    def stamp(b_row: pd.DataFrame, f_row: pd.DataFrame):
        gk = make_group_key("K4", b_row)
        b = b_row.copy(); f = f_row.copy()
        b["__MatchKategori__"]="K4"; b["__GroupKey__"]=gk
        f["__MatchKategori__"]="K4"; f["__GroupKey__"]=gk
        matched_bank_all.append(b); matched_bokf_all.append(f)

    for _, b in bank_k4.iterrows():
        b_date = pd.to_datetime(b["Bokföringsdatum"]).date() if pd.notna(b["Bokföringsdatum"]) else None
        amount = sek_round(b["Belopp"])
        if b_date is None or pd.isna(amount): continue

        # 1: exakt Datum & belopp
        amt_all = get_amount_series(bokf_df)
        cand = bokf_df[(bokf_df["Datum"].dt.date == b_date) & (~bokf_df["BokfRowID"].isin(used)) & (amt_all.round(2) == amount)].copy()
        if len(cand) >= 1:
            chosen = cand.sort_values("BokfRowID").iloc[[0]]; used |= set(chosen["BokfRowID"])
            stamp(b.to_frame().T, chosen); continue

        # 2: datum via YYMMDD i Text1
        y = extract_yymmdd(b["Bokföringsdatum"])
        mask_text = col_apply(bokf_df, "Text1", lambda t: has_yymmdd_in_text1(t, y))
        amt = get_amount_series(bokf_df)
        cand = bokf_df[mask_text & (~bokf_df["BokfRowID"].isin(used)) & (amt.round(2) == amount)].copy()
        if len(cand) >= 1:
            chosen = cand.sort_values("BokfRowID").iloc[[0]]; used |= set(chosen["BokfRowID"])
            stamp(b.to_frame().T, chosen); continue

        # 3: datum via YYMMDD i Vnr
        mask_v = col_apply(bokf_df, "Verifikationsnummer", lambda v: has_yymmdd_in_vnr(v, y))
        amt = get_amount_series(bokf_df)
        cand = bokf_df[mask_v & (~bokf_df["BokfRowID"].isin(used)) & (amt.round(2) == amount)].copy()
        if len(cand) >= 1:
            chosen = cand.sort_values("BokfRowID").iloc[[0]]; used |= set(chosen["BokfRowID"])
            stamp(b.to_frame().T, chosen); continue

    matched_bank = pd.concat(matched_bank_all, ignore_index=True) if matched_bank_all else bank_k4.iloc[0:0].copy()
    matched_bokf = pd.concat(matched_bokf_all, ignore_index=True) if matched_bokf_all else bokf_df.iloc[0:0].copy()
    return matched_bank, matched_bokf

# ===================== K4N (valuta + YYMMDD subset-sum) =====================
def run_category4N_valuta_multi(bank_df, bokf_df):
    bank_mask = bank_df["Text"].astype(str).str.contains(r"(?<!\d)3\d{5}(?!\d)", na=False, regex=True)
    bank_k4n = bank_df[bank_mask].copy().sort_values(["Bokföringsdatum","BankRowID"])
    matched_bank_all, matched_bokf_all, used = [], [], set()

    def stamp(b_rows: pd.DataFrame, f_rows: pd.DataFrame):
        gk = make_group_key("K4N", b_rows)
        b = b_rows.copy(); f = f_rows.copy()
        b["__MatchKategori__"]="K4N"; b["__GroupKey__"]=gk
        f["__MatchKategori__"]="K4N"; f["__GroupKey__"]=gk
        matched_bank_all.append(b); matched_bokf_all.append(f); return set(f["BokfRowID"])

    for _, b in bank_k4n.iterrows():
        y = extract_yymmdd(b["Bokföringsdatum"]); bank_amt = sek_round(b["Belopp"])
        cand = bokf_df[(~bokf_df["BokfRowID"].isin(used)) & (col_apply(bokf_df, "Text1", lambda t: ("valuta" in (t or "").lower()) and has_yymmdd_in_text1(t, y)))].copy()
        if cand.empty: continue
        cur = cand.copy(); amounts = get_amount_series(cur)
        hit_idx = subset_sum_meet_in_middle(cur.index.tolist(), amounts.tolist(), bank_amt, tol=0.005)
        if hit_idx:
            chosen = cur.loc[list(hit_idx)].copy()
            used |= stamp(b.to_frame().T, chosen)

    matched_bank = pd.concat(matched_bank_all, ignore_index=True) if matched_bank_all else bank_k4n.iloc[0:0].copy()
    matched_bokf = pd.concat(matched_bokf_all, ignore_index=True) if matched_bokf_all else bokf_df.iloc[0:0].copy()
    return matched_bank, matched_bokf

# ===================== K5 (LB) – inkl. steg 7 + 9–13 =====================
def run_category5_LB(bank_df: pd.DataFrame, bokf_df: pd.DataFrame):
    bank_lb = bank_df[bank_df["Text"].astype(str).str.match(r"^\s*LB", case=False, na=False)].copy()
    matched_bank_all, matched_bokf_all = [], []
    used: Set[int] = set()

    def stamp(b_rows: pd.DataFrame, f_rows: pd.DataFrame):
        gk = make_group_key("K5", b_rows)
        b = b_rows.copy(); f = f_rows.copy()
        b["__MatchKategori__"]="K5"; b["__GroupKey__"]=gk
        f["__MatchKategori__"]="K5"; f["__GroupKey__"]=gk
        matched_bank_all.append(b); matched_bokf_all.append(f); return set(f["BokfRowID"])

    def try_match(df_now: pd.DataFrame, target_sum: float) -> bool:
        return math.isclose(sum_sek(get_amount_series(df_now)), target_sum, abs_tol=0.005)

    for bank_date, bank_day_rows in bank_lb.groupby(bank_lb["Bokföringsdatum"].dt.date):
        bank_day_rows = bank_day_rows.sort_values("BankRowID")
        bank_sum = sum_sek(bank_day_rows["Belopp"])
        y = extract_yymmdd(pd.to_datetime(bank_date))

        def get_all(neg_only: bool, date_limited: bool=True):
            q = (~bokf_df["BokfRowID"].isin(used))
            if date_limited: q &= (bokf_df["Datum"].dt.date == bank_date)
            if neg_only: q &= (bokf_df["Period SEK"] < 0)
            return bokf_df[q].copy()

        # 1–3 alla samma dag
        bokf_all = get_all(neg_only=False, date_limited=True)
        if not bokf_all.empty:
            if try_match(bokf_all, bank_sum): used |= stamp(bank_day_rows, bokf_all); continue
            cand = bokf_all[get_amount_series(bokf_all).round(2) == bank_sum]
            if len(cand) >= 1: used |= stamp(bank_day_rows, cand.sort_values("BokfRowID").iloc[[0]]); continue
            diff = sek_round(sum_sek(get_amount_series(bokf_all)) - bank_sum)
            if diff != 0:
                drop = bokf_all[get_amount_series(bokf_all).round(2) == diff]
                if len(drop) >= 1:
                    drop_id = drop.sort_values("BokfRowID").iloc[0]["BokfRowID"]
                    remainder = bokf_all[bokf_all["BokfRowID"] != drop_id]
                    if try_match(remainder, bank_sum): used |= stamp(bank_day_rows, remainder); continue

        # 4–6 negativa samma dag
        bokf_neg = get_all(neg_only=True, date_limited=True)
        if not bokf_neg.empty:
            if try_match(bokf_neg, bank_sum): used |= stamp(bank_day_rows, bokf_neg); continue
            cand = bokf_neg[get_amount_series(bokf_neg).round(2) == bank_sum]
            if len(cand) >= 1: used |= stamp(bank_day_rows, cand.sort_values("BokfRowID").iloc[[0]]); continue
            diff = sek_round(sum_sek(get_amount_series(bokf_neg)) - bank_sum)
            if diff != 0:
                drop = bokf_neg[get_amount_series(bokf_neg).round(2) == diff]
                if len(drop) >= 1:
                    drop_id = drop.sort_values("BokfRowID").iloc[0]["BokfRowID"]
                    remainder = bokf_neg[bokf_neg["BokfRowID"] != drop_id]
                    if try_match(remainder, bank_sum): used |= stamp(bank_day_rows, remainder); continue

        # 7 subset-sum negativa samma dag
        if not bokf_neg.empty:
            cur2 = drop_subset_if_sum_matches(bokf_neg, bank_sum)
            if cur2 is not None and try_match(cur2, bank_sum): used |= stamp(bank_day_rows, cur2); continue

        # 8 en negativ rad == bank (oavsett datum)
        bokf_neg_any = get_all(neg_only=True, date_limited=False)
        if not bokf_neg_any.empty:
            cand = bokf_neg_any[get_amount_series(bokf_neg_any).round(2) == bank_sum]
            if len(cand) >= 1: used |= stamp(bank_day_rows, cand.sort_values("BokfRowID").iloc[[0]]); continue

        # 9–13 mix: negativa samma dag + negativa med rätt YYMMDD i Text1 (oavsett datum)
        neg_same = get_all(neg_only=True, date_limited=True)
        neg_y = bokf_df[(~bokf_df["BokfRowID"].isin(used)) & (bokf_df["Period SEK"] < 0) &
                        (col_apply(bokf_df, "Text1", lambda t: has_yymmdd_in_text1(t, y)))].copy()
        mix = pd.concat([neg_same, neg_y], ignore_index=False).drop_duplicates(subset=["BokfRowID"])

        if not mix.empty and try_match(mix, bank_sum): used |= stamp(bank_day_rows, mix); continue
        if not mix.empty:
            cand = mix[get_amount_series(mix).round(2) == bank_sum]
            if len(cand) >= 1: used |= stamp(bank_day_rows, cand.sort_values("BokfRowID").iloc[[0]]); continue
        if not mix.empty:
            diff = sek_round(sum_sek(get_amount_series(mix)) - bank_sum)
            if diff != 0:
                drop = mix[get_amount_series(mix).round(2) == diff]
                if len(drop) >= 1:
                    drop_id = drop.sort_values("BokfRowID").iloc[0]["BokfRowID"]
                    remainder = mix[mix["BokfRowID"] != drop_id]
                    if try_match(remainder, bank_sum): used |= stamp(bank_day_rows, remainder); continue

        neg_all = get_all(neg_only=True, date_limited=False)
        mix2 = pd.concat([neg_all, neg_y], ignore_index=False).drop_duplicates(subset=["BokfRowID"])
        if not mix2.empty:
            cand = mix2[get_amount_series(mix2).round(2) == bank_sum]
            if len(cand) >= 1: used |= stamp(bank_day_rows, cand.sort_values("BokfRowID").iloc[[0]]); continue
        if not mix2.empty:
            cur2 = drop_subset_if_sum_matches(mix2, bank_sum)
            if cur2 is not None and try_match(cur2, bank_sum): used |= stamp(bank_day_rows, cur2); continue

    matched_bank = pd.concat(matched_bank_all, ignore_index=True) if matched_bank_all else bank_lb.iloc[0:0].copy()
    matched_bokf = pd.concat(matched_bokf_all, ignore_index=True) if matched_bokf_all else bokf_df.iloc[0:0].copy()
    return matched_bank, matched_bokf



# ===================== K5N (Global balans per datum, ta bort minsta mängd) =====================
def run_category5N_global_balance(bank_df: pd.DataFrame, bokf_df: pd.DataFrame):
    """
    Körs per datum på kvarvarande rader.
    För varje datum: beräkna diff = bokf_sum - bank_sum.
    Försök först ta bort enstaka bokföringsrad == diff, därefter kombination (subset-sum).
    Om ej lyckat: testa banksidan motsvarande (rad == -diff, därefter kombination).
    Vid träff stämplas samtliga återstående rader för datumet som K5N.
    """
    if bank_df.empty and bokf_df.empty:
        return bank_df.iloc[0:0].copy(), bokf_df.iloc[0:0].copy()

    matched_bank_all, matched_bokf_all = [], []

    # Dagar där det finns rader på båda sidor
    bank_days = set(bank_df.dropna(subset=["Bokföringsdatum"])["Bokföringsdatum"].dt.date.unique())
    bokf_days = set(bokf_df.dropna(subset=["Datum"])["Datum"].dt.date.unique())
    common_days = sorted(bank_days & bokf_days)

    def try_stamp(day_bank, day_bokf, removed_bank_idx=None, removed_bokf_idx=None):
        """Returnera (B_rem, F_rem) om balansen håller efter att 'removed_*' tagits bort, annars None."""
        b_rem = day_bank
        f_rem = day_bokf
        if removed_bank_idx:
            b_rem = day_bank.drop(index=list(removed_bank_idx))
        if removed_bokf_idx:
            f_rem = day_bokf.drop(index=list(removed_bokf_idx))

        bank_sum = sum_sek(b_rem["Belopp"])
        bokf_sum = sum_sek(get_amount_series(f_rem))
        if math.isclose(bank_sum, bokf_sum, abs_tol=0.005):
            # Stämpla
            gk = make_group_key("K5N", b_rem if not b_rem.empty else day_bank)
            if not b_rem.empty:
                b = b_rem.copy()
                b["__MatchKategori__"] = "K5N"; b["__GroupKey__"] = gk
                matched_bank_all.append(b)
            if not f_rem.empty:
                f = f_rem.copy()
                f["__MatchKategori__"] = "K5N"; f["__GroupKey__"] = gk
                matched_bokf_all.append(f)
            return True
        return False

    for d in common_days:
        day_bank = bank_df[bank_df["Bokföringsdatum"].dt.date == d].copy()
        day_bokf = bokf_df[bokf_df["Datum"].dt.date == d].copy()
        if day_bank.empty or day_bokf.empty:  # kräver båda sidor
            continue

        bank_sum = sum_sek(day_bank["Belopp"])
        bokf_sum = sum_sek(get_amount_series(day_bokf))
        diff = sek_round(bokf_sum - bank_sum)

        # -- Bokföringssidan: singel == diff
        amt_f = get_amount_series(day_bokf).round(2)
        single_f = day_bokf[amt_f == diff]
        if not single_f.empty:
            if try_stamp(day_bank, day_bokf, removed_bokf_idx={single_f.sort_values("BokfRowID").iloc[0].name}):
                continue

        # -- Bokföringssidan: subset-sum == diff (begränsa till max 50 rader för MITM)
        candidates_f = day_bokf.copy()
        if len(candidates_f) > 50:
            candidates_f = candidates_f.reindex(candidates_f["Period SEK"].abs().sort_values(ascending=False).index[:50])
        hit = subset_sum_meet_in_middle(candidates_f.index.tolist(),
                                        get_amount_series(candidates_f).astype(float).tolist(),
                                        diff, tol=0.005)
        if hit:
            if try_stamp(day_bank, day_bokf, removed_bokf_idx=set(hit)):
                continue

        # -- Banksidan: singel == -diff
        single_b = day_bank[day_bank["Belopp"].round(2) == -diff]
        if not single_b.empty:
            if try_stamp(day_bank, day_bokf, removed_bank_idx={single_b.sort_values("BankRowID").iloc[0].name}):
                continue

        # -- Banksidan: subset-sum == -diff (max 50 rader)
        candidates_b = day_bank.copy()
        if len(candidates_b) > 50:
            candidates_b = candidates_b.reindex(candidates_b["Belopp"].abs().sort_values(ascending=False).index[:50])
        hit_b = subset_sum_meet_in_middle(candidates_b.index.tolist(),
                                          candidates_b["Belopp"].astype(float).tolist(),
                                          -diff, tol=0.005)
        if hit_b:
            if try_stamp(day_bank, day_bokf, removed_bank_idx=set(hit_b)):
                continue

    matched_bank = pd.concat(matched_bank_all, ignore_index=True) if matched_bank_all else bank_df.iloc[0:0].copy()
    matched_bokf = pd.concat(matched_bokf_all, ignore_index=True) if matched_bokf_all else bokf_df.iloc[0:0].copy()
    return matched_bank, matched_bokf
# ===================== K6 (Symmetrisk) =====================
def run_category6_symmetric(bank_df, bokf_df):
    if bank_df.empty and bokf_df.empty:
        return bank_df.iloc[0:0].copy(), bokf_df.iloc[0:0].copy()

    bank_df = bank_df.copy(); bank_df["__flip__"] = -bank_df["Belopp"]
    bank_sum = bank_df.dropna(subset=["Bokföringsdatum"]).groupby(bank_df["Bokföringsdatum"].dt.date)["__flip__"].sum().round(2)
    bokf_sum = bokf_df.dropna(subset=["Datum"]).groupby(bokf_df["Datum"].dt.date)["Period SEK"].sum().round(2)

    all_dates = sorted(set(bank_sum.index) | set(bokf_sum.index))
    totals = {d: round(float(bank_sum.get(d,0.0) + bokf_sum.get(d,0.0)), 2) for d in all_dates}
    matched_dates = {d for d,t in totals.items() if math.isclose(t, 0.0, abs_tol=0.005)}

    rem = {d:t for d,t in totals.items() if d not in matched_dates and not math.isclose(t,0.0, abs_tol=0.005)}
    plus_days  = [(d,t) for d,t in rem.items() if t > 0]
    minus_days = [(d,t) for d,t in rem.items() if t < 0]

    used_plus, used_minus, combo_groups = set(), set(), []

    def find_subset_sum(items_pos, target_pos, max_k=10, max_combos=2000):
        tried = 0; values = sorted(items_pos, key=lambda x: x[1], reverse=True)
        for r in range(1, min(max_k, len(values)) + 1):
            for combo in itertools.combinations(values, r):
                tried += 1
                if tried > max_combos: return None
                s = round(sum(v for _, v in combo), 2)
                if math.isclose(s, target_pos, abs_tol=0.005):
                    return {d for d,_ in combo}
        return None

    for d_plus, v_plus in plus_days:
        if d_plus in used_plus: continue
        cand = [(d, abs(v)) for d,v in minus_days if d not in used_minus]
        if not cand: continue
        hit = find_subset_sum(cand, v_plus)
        if hit:
            used_plus.add(d_plus); used_minus |= hit
            combo_groups.append({"dates": {d_plus, *hit}})

    for d_minus, v_minus in minus_days:
        if d_minus in used_minus: continue
        cand = [(d, v) for d,v in plus_days if d not in used_plus]
        if not cand: continue
        hit = find_subset_sum(cand, abs(v_minus))
        if hit:
            used_minus.add(d_minus); used_plus |= hit
            combo_groups.append({"dates": {d_minus, *hit}})

    # Stämpla grupper (dagar 1–1 och grupper)
    matched_bank, matched_bokf = [], []
    single_dates = sorted(d for d in totals if d in matched_dates and all(d not in g["dates"] for g in combo_groups))

    for d in single_dates:
        b_rows = bank_df[bank_df["Bokföringsdatum"].dt.date == d].copy()
        f_rows = bokf_df[bokf_df["Datum"].dt.date == d].copy()
        if not b_rows.empty or not f_rows.empty:
            gk = make_group_key("K6", b_rows)
            if not b_rows.empty: b_rows["__MatchKategori__"]="K6"; b_rows["__GroupKey__"]=gk; matched_bank.append(b_rows)
            if not f_rows.empty: f_rows["__MatchKategori__"]="K6"; f_rows["__GroupKey__"]=gk; matched_bokf.append(f_rows)

    for g in combo_groups:
        dset = g["dates"]
        b_rows = bank_df[bank_df["Bokföringsdatum"].dt.date.isin(dset)].copy()
        f_rows = bokf_df[bokf_df["Datum"].dt.date.isin(dset)].copy()
        if not b_rows.empty or not f_rows.empty:
            gk = make_group_key("K6", b_rows)
            if not b_rows.empty: b_rows["__MatchKategori__"]="K6"; b_rows["__GroupKey__"]=gk; matched_bank.append(b_rows)
            if not f_rows.empty: f_rows["__MatchKategori__"]="K6"; f_rows["__GroupKey__"]=gk; matched_bokf.append(f_rows)

    matched_bank = pd.concat(matched_bank, ignore_index=True) if matched_bank else bank_df.iloc[0:0].copy()
    matched_bokf = pd.concat(matched_bokf, ignore_index=True) if matched_bokf else bokf_df.iloc[0:0].copy()
    return matched_bank, matched_bokf

# ===================== KOMBINERAD =====================
def build_combined_all(bank_all, bokf_all, mapping_bank, mapping_bokf):
    bank_rows = []
    for _, r in bank_all.iterrows():
        is_matched = r["BankRowID"] in mapping_bank
        cat, gid = mapping_bank.get(r["BankRowID"], ("",""))
        text = str(r.get("Text","") or "")
        # Ny källa (Bank)
        if is_matched:
            ny_kalla = "Match"
        elif re.match(r"^\s*BG04803458", text, flags=re.IGNORECASE):
            ny_kalla = "Kundreskontra"
        elif re.match(r"^\s*BG04868550", text, flags=re.IGNORECASE):
            ny_kalla = "Manuell"
        elif re.match(r"^\s*LB", text, flags=re.IGNORECASE):
            ny_kalla = "Leverantörsreskontra"
        else:
            ny_kalla = "Manuell"

        row = {col:"" for col in KOMB_COLS}
        row["Datum"] = r["Bokföringsdatum"]
        row["Period SEK"] = -float(r["Belopp"]) if pd.notna(r["Belopp"]) else None
        row["Text"] = text
        row["Verifikationsnummer"] = ""
        row["System"] = "Bank"
        row["Ny källa"] = ny_kalla
        row["MatchKategori"] = cat
        row["MatchGruppID"] = gid
        bank_rows.append(row)

    bokf_rows = []
    for _, r in bokf_all.iterrows():
        is_matched = r["BokfRowID"] in mapping_bokf
        cat, gid = mapping_bokf.get(r["BokfRowID"], ("",""))
        ny_kalla = "Match" if is_matched else (r.get("Källa","") or "")

        row = {col:"" for col in KOMB_COLS}
        row["Gruppering: (KTO-ANS-SPE)"] = r.get("Gruppering: (KTO-ANS-SPE)","")
        row["FTG"] = r.get("FTG","")
        row["KTO"] = r.get("KTO","")
        row["SPE"] = r.get("SPE","")
        row["ANS"] = r.get("ANS","")
        row["OBJ"] = r.get("OBJ","")
        row["MOT"] = r.get("MOT","")
        row["PRD"] = r.get("PRD","")
        row["MAR"] = r.get("MAR","")
        row["RGR"] = r.get("RGR","")
        row["Datum"] = r.get("Datum","")
        row["IB Året SEK"] = r.get("IB Året SEK","")
        row["Ing. ack. Belopp"] = r.get("Ing. ack. belopp 07-2025 SEK","")
        row["Period SEK"] = r.get("Period SEK","")
        row["Utg. ack. Belopp"] = r.get("Utg. ack. belopp 07-2025 SEK","")
        row["Val"] = r.get("Val","")
        row["Utländskt valutabelopp"] = r.get("Utländskt valutabelopap","")
        row["Text"] = r.get("Text1","")
        row["Postning -Dokumentsekvensnummer"] = r.get("Postning -Dokumentsekvensnummer","")
        row["Verifikationsnummer"] = r.get("Verifikationsnummer","")
        row["Källa"] = r.get("Källa","")
        row["Kategori"] = r.get("Kategori","")
        row["System"] = "Bokföring"
        row["Ny källa"] = ny_kalla
        row["MatchKategori"] = cat
        row["MatchGruppID"] = gid
        bokf_rows.append(row)

    komb = pd.DataFrame(bank_rows + bokf_rows, columns=KOMB_COLS)
    komb["System"] = komb["System"].astype(pd.CategoricalDtype(["Bank","Bokföring"], ordered=True))
    komb = komb.sort_values(by=["MatchGruppID","Datum","System"], na_position="last").reset_index(drop=True)
    return komb

def _argb(hex_rgb: str) -> str:
    rgb = (hex_rgb or "").lstrip("#").upper()
    if len(rgb) != 6: rgb = "FFFFFF"
    return "FF" + rgb

def make_combined_sheet(wb_path: Path):
    wb = load_workbook(wb_path)
    ws = wb["Kombinerad"]

    def fill(cell, value=None, bg_hex=None, border=True, bold=False):
        if value is not None: ws[cell] = value
        if bg_hex:
            argb = _argb(bg_hex)
            ws[cell].fill = PatternFill(start_color=argb, end_color=argb, fill_type="solid")
        if border:
            thin = Side(style="thin", color="000000")
            ws[cell].border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws[cell].alignment = Alignment(vertical="center")
        if bold: ws[cell].font = Font(bold=True)

    # Rad 2 – etiketter/format
    fill("B2", "Bank", bg_hex="#B8D3EF", bold=True)
    fill("C2", "", border=True); ws["C2"].number_format = "#,##0.00"
    fill("D2", "Bokföring", bg_hex="#B8D3EF", bold=True)
    fill("E2", "", border=True); ws["E2"].number_format = "#,##0.00"
    fill("G2", bg_hex="#D9D9D9", border=True); ws["G2"] = "=E2-C2"; ws["G2"].number_format = "#,##0.00"
    fill("N2", bg_hex="#D9D9D9", border=True); ws["N2"] = "=ROUND(SUBTOTAL(9,N5:N99999),2)"; ws["N2"].number_format = "#,##0.00"

    # Frys rubriker (rad 4)
    ws.freeze_panes = "A5"

    # Rubrikrad (rad 4)
    for col_idx in range(1, ws.max_column + 1):
        cell = f"{get_column_letter(col_idx)}4"
        fill(cell, ws[cell].value, bg_hex="#B8D3EF", border=True, bold=True)

    # Kolumnbredder
    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    # N (Period SEK) från rad 5
    col_N = 14
    for r in range(5, ws.max_row + 1):
        ws.cell(row=r, column=col_N).number_format = "#,##0.00"

    # K (Datum) från rad 5
    col_K = 11
    for r in range(5, ws.max_row + 1):
        ws.cell(row=r, column=col_K).number_format = "yyyy-mm-dd"

    # X (Ny källa) grå från rad 5
    gray = _argb("#D9D9D9")
    for r in range(5, ws.max_row + 1):
        ws.cell(row=r, column=24).fill = PatternFill(start_color=gray, end_color=gray, fill_type="solid")

    # AutoFilter
    last_col_letter = get_column_letter(ws.max_column)
    last_row = max(4, ws.max_row)
    ws.auto_filter.ref = f"A4:{last_col_letter}{last_row}"

    wb.save(wb_path)

# ===================== EXPORT =====================
def drop_ids(df: pd.DataFrame) -> pd.DataFrame:
    return df.drop(columns=["BankRowID","BokfRowID"], errors="ignore")

def df_by_cat(df: pd.DataFrame, cat: str) -> pd.DataFrame:
    if df.empty or "__MatchKategori__" not in df.columns: return df.iloc[0:0].copy()
    return df[df["__MatchKategori__"] == cat].copy()


def _write_result_excel(out_path: str, komb: pd.DataFrame) -> str:
    komb = clean_for_excel(komb)

    out = Path(out_path)
    with pd.ExcelWriter(out, engine="openpyxl") as xw:
        komb.to_excel(xw, index=False, sheet_name="Kombinerad", startrow=3)

    make_combined_sheet(out)

    try:
        with zipfile.ZipFile(out, "r") as zf:
            _ = zf.namelist()
    except zipfile.BadZipFile:
        raise RuntimeError("XLSX blev korrupt – sannolikt låsning eller icke-sanerat värde.")
    return str(out)

# ===================== MAIN =====================
def main():
    print("🔹 Först väljer du kontoutdraget.\n🔹 Sen väljer du bokföringslistan.\n")
    bank_path = pick_file_with_validation("Bank")
    bokf_path = pick_file_with_validation("Bokföring")

    initialdir = str(Path(bank_path).parent) if bank_path else None
    out_path = ask_save_as_dialog("output_avstamning.xlsx", initialdir=initialdir)
    if not out_path:
        print("Ingen sparfil vald – avbryter."); return

    bank_all = load_bank(bank_path)
    bokf_all = load_bokf(bokf_path)

    bank_rem = bank_all.copy()
    bokf_rem = bokf_all.copy()
    matched_bank_list, matched_bokf_list = [], []

    for cat, func in [("K1",run_category1_BG04803458),
                      ("K2",run_category2_BG04868550),
                      ("K3",run_category3_35ref),
                      ("K4",run_category4_ovrigt),
                      ("K4N",run_category4N_valuta_multi),
                      ("K5",run_category5_LB),
                      ("K5N",run_category5N_global_balance)]:
        mb, mf = func(bank_rem, bokf_rem)
        if not mb.empty: mb = mb.copy(); mb["__MatchKategori__"] = mb.get("__MatchKategori__","K?")  # redan satt i stamp
        if not mf.empty: mf = mf.copy(); mf["__MatchKategori__"] = mf.get("__MatchKategori__","K?")
        matched_bank_list.append(mb); matched_bokf_list.append(mf)
        if not mb.empty: bank_rem = bank_rem[~bank_rem["BankRowID"].isin(mb["BankRowID"])]
        if not mf.empty: bokf_rem = bokf_rem[~bokf_rem["BokfRowID"].isin(mf["BokfRowID"])]

    # K6 på rester
    mb6, mf6 = run_category6_symmetric(bank_rem, bokf_rem)
    if not mb6.empty: mb6["__MatchKategori__"] = "K6"
    if not mf6.empty: mf6["__MatchKategori__"] = "K6"
    matched_bank_list.append(mb6); matched_bokf_list.append(mf6)

    matched_bank_all = pd.concat([d for d in matched_bank_list if not d.empty], ignore_index=True) if any((not d.empty for d in matched_bank_list)) else bank_all.iloc[0:0].copy()
    matched_bokf_all = pd.concat([d for d in matched_bokf_list if not d.empty], ignore_index=True) if any((not d.empty for d in matched_bokf_list)) else bokf_all.iloc[0:0].copy()

    if "__MatchKategori__" not in matched_bank_all.columns: matched_bank_all["__MatchKategori__"] = pd.Series(dtype=str)
    if "__MatchKategori__" not in matched_bokf_all.columns: matched_bokf_all["__MatchKategori__"] = pd.Series(dtype=str)

    om_bank_all = bank_all[~bank_all["BankRowID"].isin(matched_bank_all.get("BankRowID", pd.Series(dtype=int)))].copy()
    om_bokf_all = bokf_all[~bokf_all["BokfRowID"].isin(matched_bokf_all.get("BokfRowID", pd.Series(dtype=int)))].copy()

    # Bygg mapping ENBART från __GroupKey__
    mapping_bank, mapping_bokf = {}, {}
    if "__GroupKey__" in matched_bank_all.columns:
        for g, bgrp in matched_bank_all.dropna(subset=["__GroupKey__"]).groupby("__GroupKey__"):
            cat = (bgrp["__MatchKategori__"].dropna().iloc[0]) if "__MatchKategori__" in bgrp.columns and not bgrp["__MatchKategori__"].dropna().empty else ""
            for bid in bgrp.get("BankRowID", []): mapping_bank[bid] = (cat, g)
    if "__GroupKey__" in matched_bokf_all.columns:
        for g, fgrp in matched_bokf_all.dropna(subset=["__GroupKey__"]).groupby("__GroupKey__"):
            cat = (fgrp["__MatchKategori__"].dropna().iloc[0]) if "__MatchKategori__" in fgrp.columns and not fgrp["__MatchKategori__"].dropna().empty else ""
            for fid in fgrp.get("BokfRowID", []): mapping_bokf[fid] = (cat, g)

    komb = build_combined_all(bank_all, bokf_all, mapping_bank, mapping_bokf)

    out_written = _write_result_excel(out_path, komb)
    print(f"✅ Klar! Skrev: {out_written}")

if __name__ == "__main__":
    main()
